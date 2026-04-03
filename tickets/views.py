"""
tickets/views.py
================
Django views — the Controller layer in MVC.

New in this version:
  - employee_login / employee_logout — session-based authentication
  - login_required decorator — guards submit, success, and routing pages
  - submit_ticket pre-fills form from logged-in employee's profile
  - export_tickets — generates CSV, XLSX, or PDF of the current queue
"""

import csv
import io
import json
from functools import wraps

from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponse

from .forms import TicketSubmitForm
from .models import Employee, EmployeeRole, Ticket, TicketStatus, TicketHistory, HistoryAction
from .routing import (
    RoutingEngine,
    TIER_META,
    SLA_MATRIX,
    CATEGORY_TEAMS,
    SUBTYPE_RULES,
    KEYWORD_RULES,
    DEPARTMENT_TIERS,
    ISSUE_CASCADE,
    PRIORITY_LABELS,
)


# ---------------------------------------------------------------------------
# AUTH HELPERS
# ---------------------------------------------------------------------------

def login_required(view_fn):
    """Redirect unauthenticated users to the login page."""
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.session.get("employee_pk"):
            return redirect("employee_login")
        return view_fn(request, *args, **kwargs)
    return wrapper


def ist_required(view_fn):
    """Restrict a view to IST staff and admins only.
    Authenticated dept employees are redirected to the submit page."""
    @wraps(view_fn)
    def wrapper(request, *args, **kwargs):
        if not request.session.get("employee_pk"):
            return redirect("employee_login")
        role = request.session.get("employee_role", "dept")
        if role not in ("ist", "admin"):
            return redirect("submit_ticket")
        return view_fn(request, *args, **kwargs)
    return wrapper


def _get_logged_in_employee(request):
    """Return the Employee object for the current session, or None."""
    emp_pk = request.session.get("employee_pk")
    if emp_pk:
        try:
            return Employee.objects.get(pk=emp_pk, is_active=True)
        except Employee.DoesNotExist:
            pass
    return None


# ---------------------------------------------------------------------------
# EMPLOYEE LOGIN / LOGOUT
# ---------------------------------------------------------------------------

def employee_login(request):
    """
    GET  — show login form.
    POST — validate employee_id + email + password against Employee table.
    """
    # Already logged in → go to submit form
    if request.session.get("employee_pk"):
        return redirect("submit_ticket")

    error = ""

    if request.method == "POST":
        emp_id   = request.POST.get("employee_id", "").strip()
        email    = request.POST.get("email", "").strip().lower()
        password = request.POST.get("password", "")

        try:
            emp = Employee.objects.get(employee_id=emp_id, is_active=True)
            if emp.email.lower() == email and emp.check_password(password):
                # Store minimal session data
                request.session["employee_pk"]   = emp.pk
                request.session["employee_id"]   = emp.employee_id
                request.session["employee_name"] = emp.full_name
                request.session["employee_role"] = emp.role
                request.session.set_expiry(28800)   # 8-hour session
                return redirect("submit_ticket")
            else:
                error = "Invalid Employee ID, email, or password."
        except Employee.DoesNotExist:
            error = "Invalid Employee ID, email, or password."

    return render(request, "tickets/login.html", {"error": error})


def employee_logout(request):
    """Clear the session and redirect to login."""
    request.session.flush()
    return redirect("employee_login")


# ---------------------------------------------------------------------------
# SUBMIT TICKET
# ---------------------------------------------------------------------------

@login_required
def submit_ticket(request):
    """
    GET  — display form pre-filled from the employee's profile.
    POST — validate, run routing engine, persist, redirect to success.
    """
    emp = _get_logged_in_employee(request)

    if request.method == "POST":
        form = TicketSubmitForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.submitter = emp

            combined_text = (
                form.cleaned_data.get("title", "") + " " +
                form.cleaned_data.get("description", "")
            )
            result = RoutingEngine.compute(
                dept=form.cleaned_data["department"],
                category=form.cleaned_data["category"],
                subtype=form.cleaned_data.get("subtype", ""),
                user_priority=3,   # default Moderate; engine adjusts from here
                text=combined_text,
            )

            ticket.routing_tier               = result.tier
            ticket.routing_tier_label         = result.tier_label
            ticket.routing_team               = result.team
            ticket.routing_sla                = result.sla
            ticket.routing_effective_priority = result.effective_priority
            ticket.routing_was_modified       = result.was_modified
            ticket.routing_reasons            = json.dumps(result.reasons)
            ticket.routing_escalation_path    = json.dumps(result.escalation_path)

            ticket.save()

            # Log the creation event
            TicketHistory.objects.create(
                ticket=ticket,
                action=HistoryAction.CREATED,
                changed_by=ticket.name,
                team_after=ticket.routing_team,
                priority_after=ticket.routing_effective_priority,
                note="Ticket submitted via self-service portal.",
            )

            return redirect("ticket_success", ticket_id=ticket.ticket_id)
    else:
        # Pre-fill name / employee_id / department / email from Employee profile
        initial = {}
        if emp:
            initial = {
                "name":        emp.full_name,
                "employee_id": emp.employee_id,
                "department":  emp.department,
                "email":       emp.email,
            }
        form = TicketSubmitForm(initial=initial)

    cascade_json = json.dumps({
        cat: {
            "subtypes":    data["subtypes"],
            "issue_types": data["issue_types"],
        }
        for cat, data in ISSUE_CASCADE.items()
    })

    return render(request, "tickets/submit.html", {
        "form":         form,
        "cascade_json": cascade_json,
        "employee":     emp,
    })


@login_required
def ticket_success(request, ticket_id):
    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)
    return render(request, "tickets/success.html", {"ticket": ticket})


# ---------------------------------------------------------------------------
# LIVE ROUTING AJAX ENDPOINT
# ---------------------------------------------------------------------------

@login_required
def live_route(request):
    dept     = request.GET.get("dept", "")
    category = request.GET.get("category", "")
    subtype  = request.GET.get("subtype", "")
    priority = int(request.GET.get("priority", 4))
    text     = request.GET.get("text", "")

    if not dept or not category:
        return JsonResponse({"ready": False})

    result = RoutingEngine.compute(dept, category, subtype, priority, text)

    return JsonResponse({
        "ready":              True,
        "team":               result.team,
        "sla":                result.sla,
        "effective_priority": result.effective_priority,
        "priority_label":     result.priority_label,
        "tier":               result.tier,
        "tier_label":         result.tier_label,
        "tier_icon":          result.tier_icon,
        "suggested_priority": result.suggested_priority,
        "was_modified":       result.was_modified,
        "escalation_path":    result.escalation_path,
        "reasons":            result.reasons,
    })


# ---------------------------------------------------------------------------
# ADMIN QUEUE
# ---------------------------------------------------------------------------

@ist_required
def admin_queue(request):
    """Admin dashboard — no login guard so IT staff can access directly."""
    tickets = Ticket.objects.select_related("submitter").all()

    status_filter   = request.GET.get("status",   "")
    priority_filter = request.GET.get("priority", "")
    tier_filter     = request.GET.get("tier",     "")

    if status_filter:
        tickets = tickets.filter(status=status_filter)
    if priority_filter:
        tickets = tickets.filter(routing_effective_priority=int(priority_filter))
    if tier_filter:
        tickets = tickets.filter(routing_tier=tier_filter)

    all_tickets = Ticket.objects.all()
    stats = {
        "total":       all_tickets.count(),
        "open":        all_tickets.filter(status=TicketStatus.OPEN).count(),
        "in_progress": all_tickets.filter(status=TicketStatus.IN_PROGRESS).count(),
        "critical":    all_tickets.filter(routing_effective_priority=1).count(),
        "infra":       all_tickets.filter(routing_tier="CRITICAL_INFRA").count(),
    }

    context = {
        "tickets":          tickets,
        "stats":            stats,
        "status_filter":    status_filter,
        "priority_filter":  priority_filter,
        "tier_filter":      tier_filter,
        "status_choices":   [("", "All Statuses")] + list(TicketStatus.choices),
        "priority_choices": [
            ("", "All Priorities"),
            ("1", "Critical"), ("2", "High"), ("3", "Medium"), ("4", "Low"),
        ],
        "tier_choices": [
            ("", "All Dept Tiers"),
            ("CRITICAL_INFRA", "Critical Infra"),
            ("EXECUTIVE",      "Executive"),
            ("PUBLIC_SAFETY",  "Public Safety"),
            ("STANDARD",       "Standard"),
        ],
    }
    return render(request, "tickets/admin_queue.html", context)


# ---------------------------------------------------------------------------
# TICKET LOOKUP — employee read-only view
# ---------------------------------------------------------------------------

def ticket_lookup(request):
    """
    Read-only ticket status page for employees.
    GET with no params  — show the lookup form.
    GET ?ticket_id=TKT-XXXX — find the ticket and display it.
    Employees cannot take any action; this view only reads.
    """
    ticket = None
    error  = None
    query  = request.GET.get("ticket_id", "").strip().upper()

    if query:
        try:
            ticket = Ticket.objects.get(ticket_id=query)
        except Ticket.DoesNotExist:
            error = f'No ticket found with ID "{query}". Please check the reference number and try again.'

    return render(request, "tickets/ticket_lookup.html", {
        "ticket": ticket,
        "error":  error,
        "query":  query,
    })


# ---------------------------------------------------------------------------
# TICKET DETAIL
# ---------------------------------------------------------------------------

@ist_required
def ticket_detail(request, ticket_id):
    ticket = get_object_or_404(Ticket, ticket_id=ticket_id)

    if request.method == "POST":
        action     = request.POST.get("action")
        changed_by = request.session.get("employee_name", "IST Staff")

        if action == "escalate" and ticket.routing_effective_priority > 1:
            pri_before = ticket.routing_effective_priority
            team_before = ticket.routing_team
            new_pri    = pri_before - 1
            result     = RoutingEngine.compute(
                dept=ticket.department,
                category=ticket.category,
                subtype=ticket.subtype,
                user_priority=new_pri,
                text=ticket.title + " " + ticket.description,
            )
            ticket.routing_effective_priority = result.effective_priority
            ticket.routing_team               = result.team
            ticket.routing_sla                = result.sla
            ticket.routing_reasons            = json.dumps(result.reasons)
            ticket.routing_escalation_path    = json.dumps(result.escalation_path)
            ticket.routing_was_modified       = True
            ticket.status = TicketStatus.IN_PROGRESS
            ticket.save()

            TicketHistory.objects.create(
                ticket=ticket,
                action=HistoryAction.ESCALATED,
                changed_by=changed_by,
                priority_before=pri_before,
                priority_after=result.effective_priority,
                team_before=team_before,
                team_after=result.team,
            )

        elif action == "resolve":
            ticket.status = TicketStatus.CLOSED
            ticket.save()

            TicketHistory.objects.create(
                ticket=ticket,
                action=HistoryAction.RESOLVED,
                changed_by=changed_by,
            )

        elif action == "add_note":
            note_text = request.POST.get("note_text", "").strip()
            if note_text:
                TicketHistory.objects.create(
                    ticket=ticket,
                    action=HistoryAction.NOTE,
                    changed_by=changed_by,
                    note=note_text,
                )

        return redirect("ticket_detail", ticket_id=ticket_id)

    history = ticket.history.all()
    return render(request, "tickets/ticket_detail.html", {
        "ticket":  ticket,
        "history": history,
    })


# ---------------------------------------------------------------------------
# EXPORT REPORTS
# ---------------------------------------------------------------------------

def _get_export_queryset(request):
    """Apply same filters as admin_queue, return a queryset."""
    qs = Ticket.objects.select_related("submitter").all()
    if request.GET.get("status"):
        qs = qs.filter(status=request.GET["status"])
    if request.GET.get("priority"):
        qs = qs.filter(routing_effective_priority=int(request.GET["priority"]))
    if request.GET.get("tier"):
        qs = qs.filter(routing_tier=request.GET["tier"])
    return qs


EXPORT_COLUMNS = [
    ("ticket_id",               "Ticket ID"),
    ("submitted_at",            "Submitted"),
    ("name",                    "Employee Name"),
    ("employee_id",             "Employee ID"),
    ("department",              "Department"),
    ("routing_tier_label",      "Dept Tier"),
    ("category",                "Category"),
    ("subtype",                 "Sub-Type"),
    ("issue_type",              "Issue Type"),
    ("title",                   "Summary"),
    ("routing_effective_priority", "Eff. Priority"),
    ("routing_team",            "Assigned Team"),
    ("routing_sla",             "SLA Target"),
    ("status",                  "Status"),
    ("routing_was_modified",    "Auto-Adjusted"),
]


def _ticket_row(ticket):
    return [
        ticket.ticket_id,
        ticket.submitted_at.strftime("%Y-%m-%d %H:%M"),
        ticket.name,
        ticket.employee_id,
        ticket.department,
        ticket.routing_tier_label,
        ticket.category,
        ticket.subtype,
        ticket.issue_type,
        ticket.title,
        PRIORITY_LABELS.get(ticket.routing_effective_priority, ""),
        ticket.routing_team,
        ticket.routing_sla,
        ticket.status,
        "Yes" if ticket.routing_was_modified else "No",
    ]


@ist_required
def export_csv(request):
    tickets = _get_export_queryset(request)
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="laredo_ist_tickets.csv"'

    writer = csv.writer(response)
    writer.writerow([col[1] for col in EXPORT_COLUMNS])
    for ticket in tickets:
        writer.writerow(_ticket_row(ticket))

    return response


@ist_required
def export_xlsx(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    tickets = _get_export_queryset(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Work Orders"

    # ── Colour palette
    NAVY   = "08111F"
    GOLD   = "C9A84C"
    WHITE  = "EEF2F8"
    GREY   = "1E3358"

    # ── Title row
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = "City of Laredo IST — Work Order Export"
    title_cell.font      = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill      = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row
    header_fill   = PatternFill("solid", fgColor=GREY)
    header_font   = Font(name="Calibri", bold=True, size=10, color=GOLD)
    header_border = Border(
        bottom=Side(style="thin", color=GOLD),
        right=Side(style="thin", color="2A3A55"),
    )

    for col_idx, (_, label) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.border    = header_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 32

    # ── Data rows — zebra striping
    PRIORITY_COLOURS = {"Critical": "3D0C0C", "High": "3D2700", "Medium": "0E1E35", "Low": "111B2B"}
    EVEN_FILL = PatternFill("solid", fgColor="0F1E35")
    ODD_FILL  = PatternFill("solid", fgColor="162845")
    data_font = Font(name="Calibri", size=10, color=WHITE)

    for row_idx, ticket in enumerate(tickets, start=3):
        row_data = _ticket_row(ticket)
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font      = data_font
            cell.fill      = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Highlight priority cell
        pri_label = row_data[10]   # column index of Eff. Priority
        pri_cell  = ws.cell(row=row_idx, column=11)
        pri_colour = PRIORITY_COLOURS.get(pri_label, "111B2B")
        pri_cell.fill = PatternFill("solid", fgColor=pri_colour)
        pri_cell.font = Font(name="Calibri", bold=True, size=10, color=GOLD)

    # ── Column widths
    col_widths = [12, 16, 22, 14, 22, 20, 12, 16, 18, 40, 14, 28, 10, 12, 14]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze header rows
    ws.freeze_panes = "A3"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="laredo_ist_tickets.xlsx"'
    return response


@ist_required
def export_pdf(request):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph,
            Spacer, HRFlowable,
        )
    except ImportError:
        return HttpResponse("reportlab not installed. Run: pip install reportlab", status=500)

    from django.utils import timezone as tz

    tickets = list(_get_export_queryset(request))

    # ── Page setup
    # Landscape letter = 11" x 8.5".  Margins 0.4" each side → 10.2" usable.
    PAGE       = landscape(letter)
    LEFT_MARGIN = 0.4 * inch
    USABLE_W   = PAGE[0] - (LEFT_MARGIN * 2)   # 10.2 inches

    buffer = io.BytesIO()
    doc    = SimpleDocTemplate(
        buffer,
        pagesize=PAGE,
        leftMargin=LEFT_MARGIN, rightMargin=LEFT_MARGIN,
        topMargin=0.55*inch,    bottomMargin=0.45*inch,
    )

    # ── Colours (black and white)
    BLACK     = colors.black
    WHITE     = colors.white
    DARK_GREY = colors.HexColor("#222222")
    MID_GREY  = colors.HexColor("#555555")
    LIGHT_GREY= colors.HexColor("#DDDDDD")
    HDR_BG    = colors.HexColor("#222222")
    ROW_A     = colors.white
    ROW_B     = colors.HexColor("#F5F5F5")

    PRI_LABELS_BOLD = {"Critical": "CRITICAL", "High": "HIGH", "Medium": "MEDIUM", "Low": "LOW"}

    # ── Paragraph styles for wrapping cell text
    styles = getSampleStyleSheet()

    cell_style = ParagraphStyle(
        "cell", parent=styles["Normal"],
        fontSize=7.5, leading=10,
        textColor=DARK_GREY, fontName="Helvetica",
        wordWrap="LTR", splitLongWords=True,
    )
    cell_bold = ParagraphStyle(
        "cell_bold", parent=cell_style,
        fontName="Helvetica-Bold", textColor=BLACK,
    )
    cell_mono = ParagraphStyle(
        "cell_mono", parent=cell_style,
        fontName="Courier", fontSize=7, textColor=MID_GREY,
    )
    cell_grey = ParagraphStyle(
        "cell_grey", parent=cell_style,
        textColor=MID_GREY,
    )
    hdr_style = ParagraphStyle(
        "hdr", parent=styles["Normal"],
        fontSize=7.5, leading=10,
        textColor=WHITE, fontName="Helvetica-Bold",
        alignment=1,
    )
    title_style = ParagraphStyle(
        "title", parent=styles["Heading1"],
        textColor=WHITE, backColor=HDR_BG,
        fontSize=14, spaceAfter=4, spaceBefore=0,
        leftIndent=6,
    )
    sub_style = ParagraphStyle(
        "sub", parent=styles["Normal"],
        textColor=MID_GREY, fontSize=9, spaceAfter=8,
    )

    story = []
    story.append(Paragraph("City of Laredo — IST Work Order Report", title_style))
    story.append(Paragraph(
        f"Generated: {tz.localtime().strftime('%B %d, %Y at %I:%M %p')}  ·  "
        f"Total records: {len(tickets)}",
        sub_style,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=DARK_GREY, spaceAfter=6))

    # ── Column definitions
    # Total must equal USABLE_W (10.2").
    # Widths in inches:  ID   Date   Employee  Department  Cat   Sub-Type  Summary  Priority  Team   SLA   Status
    col_w = [inch * w for w in [0.62, 0.68,   1.05,       1.40,   0.62,  0.72,     1.90,    0.65,  1.30,  0.48, 0.78]]
    # Sum = 10.20" ✓

    headers = ["ID", "Date", "Employee", "Department", "Category",
               "Sub-Type", "Summary", "Priority", "Team", "SLA", "Status"]

    table_data = [[Paragraph(h, hdr_style) for h in headers]]

    for ticket in tickets:
        pri_label = PRIORITY_LABELS.get(ticket.routing_effective_priority, "")
        table_data.append([
            Paragraph(ticket.ticket_id,                         cell_mono),
            Paragraph(ticket.submitted_at.strftime("%m/%d/%Y"), cell_grey),
            Paragraph(ticket.name,                              cell_style),
            Paragraph(ticket.department,                        cell_style),   # no truncation
            Paragraph(ticket.category.capitalize(),             cell_style),
            Paragraph(ticket.subtype or "—",                    cell_grey),
            Paragraph(ticket.title,                             cell_style),   # no truncation
            Paragraph(pri_label,                                cell_bold),
            Paragraph(ticket.routing_team,                      cell_style),   # no truncation
            Paragraph(ticket.routing_sla,                       cell_mono),
            Paragraph(ticket.status,                            cell_style),   # full word
        ])

    tbl = Table(table_data, colWidths=col_w, repeatRows=1)

    tbl_style = TableStyle([
        # Header row
        ("BACKGROUND",    (0, 0), (-1, 0),  HDR_BG),
        ("TOPPADDING",    (0, 0), (-1, 0),  7),
        ("BOTTOMPADDING", (0, 0), (-1, 0),  7),
        ("LEFTPADDING",   (0, 0), (-1, 0),  4),
        ("RIGHTPADDING",  (0, 0), (-1, 0),  4),
        # Data rows — alternating white / light grey
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [ROW_A, ROW_B]),
        ("TOPPADDING",    (0, 1), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
        ("LEFTPADDING",   (0, 1), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 1), (-1, -1), 4),
        # Grid
        ("GRID",   (0, 0), (-1, -1), 0.35, LIGHT_GREY),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ])
    # No priority background colouring — B&W only

    tbl.setStyle(tbl_style)
    story.append(tbl)

    # ── Page footer
    def on_page(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(HDR_BG)
        canvas.rect(0, 0, PAGE[0], 0.38*inch, fill=True, stroke=False)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(WHITE)
        canvas.drawString(0.4*inch, 0.13*inch,
                          "City of Laredo — Information Systems & Technology")
        canvas.drawRightString(PAGE[0] - 0.4*inch, 0.13*inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="laredo_ist_tickets.pdf"'
    return response